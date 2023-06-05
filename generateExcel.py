import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image


def main(key, screenshot):
    # Load the workbook or create a new one if it doesn't exist
    try:
        book = openpyxl.load_workbook("result.xlsx")
    except FileNotFoundError:
        book = Workbook()

    # Select the worksheet
    sheet_name = "Sheet2"
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        sheet = book.create_sheet(sheet_name)

    # Append new row to the worksheet
    img = Image(f"screenshots/{screenshot}")
    img.width = img.width * 0.3  # Adjust image width if needed
    img.height = img.height * 0.3  # Adjust image height if needed
    img.anchor = sheet.cell(row=sheet.max_row + 1, column=2).coordinate
    sheet.row_dimensions[
        sheet.max_row
    ].height = img.height  # Adjust row height to fit the image
    sheet.column_dimensions["B"].width = (
        img.width / 6
    )  # Adjust column width to fit the image

    cell_key = sheet.cell(row=sheet.max_row, column=1)
    cell_key.value = key

    sheet.add_image(img)

    # Save the workbook
    book.save("result.xlsx")
